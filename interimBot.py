#!/usr/bin/env python
# -*- coding: utf-8 -*-
## ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ ##
##																				   ##
##  ----  ----  ----	  BOT TELEGRAM SUIVI FICHES DE PAIE	  ----  ----  ----  ##
##																				   ##
## ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

## programme pour suivre via Telegram les missions d'interim et faciliter la vérification des fiches de paye.

## ~~~~~~~~~~~~~~~~~~~~~~~~~~		PARAMETRES		 ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# modules complémentaires
import os
import sys
from re import compile as reCompile
from telegram import ReplyKeyboardMarkup, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler
from telegram.ext import ConversationHandler, MessageHandler, Filters, MessageFilter
import sqlite3
from datetime import datetime as dt
from locale import setlocale, LC_ALL
from openpyxl import Workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# les erreurs critiques
class Exit(Exception):
	pass

# mise du programme en français pour les affichage de strftime
setlocale(LC_ALL, 'fr_FR.utf8')

# dossiers racine du projet
BASEPATH = os.path.dirname(os.path.realpath(sys.argv[0]))
# chemin vers la base de donnée
BDD_PATH = os.path.join(BASEPATH, "data.db")
# table de la base de donnée
BDD_TABLE = "missions"

# configuration du .env
REGEX_TOKEN = reCompile("token=[0-9]{8,10}:[a-zA-Z0-9_-]{35}")
REGEX_SVR_NAME = reCompile("server_name=[a-zA-Z0-9.-]+")
REGEX_SVR_PORT = reCompile("server_port=[0-9]+")
REGEX_MAIL_FROM = reCompile("mail_from=[a-zA-Z0-9-.@]+")
REGEX_MAIL_MDP = reCompile("mail_mdp=[a-zA-Z0-9-+/_.:;,|!%$*]+")
REGEX_MAIL_TO = reCompile("mail_to=[a-zA-Z0-9-.@]+")

# les demandes pour la création d'un nouvel enregistrement
AGENCE, DATE, LIEU, H_DEBUT, H_FIN = range(5)
# le buffer pour enregistrer les infos
TO_SAVE = []
# la colonne maxi dans laquelle on va écrire les données sur excel (lié au modèle de suivi des fiches de paies)
MAX_COL = 6


## ~~~~~~~~~~~~~~~~~~~~~~~~~~	  GESTION DU SQL	   ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# la classe qui va contenir la base de donnée
class obj_bdd():
	# fonction d'initialisation et de fermeture de la connection
	def __init__(self, FULLPATH, tableName):
		try:
			# curseur et connection de la base de donnée
			self._conn = sqlite3.connect(FULLPATH)
			self._cursor = self._conn.cursor()
			# vérification du nom de la table
			self.cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
			self.tableName = tableName
			listeTable = [k[0] for k in self.cursor.fetchall()]
			# si la table n'existe pas, on la crée
			if self.tableName not in listeTable:
				self.cursor.execute(f"CREATE TABLE IF NOT EXISTS '{BDD_TABLE}' ('agence' TEXT, 'date' TEXT PRIMARY KEY, 'lieu' TEXT, 'heure_debut' TEXT, 'heure_fin' TEXT)")
			# enregistrement de la clef primaire
			self.primaryKey = None
			self.cursor.execute(f"PRAGMA table_info({self.tableName})")
			for k in self.cursor.fetchall():
				if k[-1]:
					self.primaryKey = k[1]
					self.primaryKeyIndex = k[0]
					break
			if self.primaryKey is None:
				raise Exit(f"[!] la table '{self.tableName}' de la base de données '{FULLPATH}' n'a pas de clef primaire")
		# le chemin spécifié ne renvois vers rien
		except sqlite3.OperationalError:
			raise Exit(f"[!] la base de donnée '{FULLPATH}' est introuvable") # jamais trigger car connect crée automatiquement un fichier

	# interaction possible avec un 'with'
	def __enter__(self):
		return self

	# interaction possible avec un 'with'
	def __exit__(self, exc_type, exc_val, exc_tb):
		self.save()
		self.close()

	# interaction entre les variables privée et les "getters"
	@property
	def connection(self):
		return self._conn
	@property
	def cursor(self):
		return self._cursor

	# recuperer les infos pour une entrée de clef primaire donnée. Si c'est "all", renvoit la totalité des données de la table
	def getDatas(self, key):
		if key == "all":
			self.cursor.execute(f"SELECT * FROM {self.tableName} ORDER BY {self.primaryKey} ASC")
			return self.cursor.fetchall()
		else:
			self.cursor.execute(f"SELECT * FROM {self.tableName} WHERE {self.primaryKey} LIKE '{key}'")
			return self.cursor.fetchone()

	# récupere les noms des champs de la table
	def namesColonnes(self):
		self.cursor.execute(f"PRAGMA table_info({self.tableName})")
		L = [k[1] for k in self.cursor.fetchall()]
		return L

	# renvois True si l'entrée de la clef primaire est bien présente dans la table
	def verify(self, key, prefixe, suffixe):
		# si prefixe et suffixe valent False, la clef doit exactement etre présente
		if not prefixe and not suffixe:
			self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '{key}'")
		# si seul prefixe vaut True, la clef doit seulement commencer pareil
		elif prefixe and not suffixe:
			self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '{key}%'")
		# si seul suffixe vaut True, la clef doit seulement finir pareil
		elif not prefixe and suffixe:
			self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '%{key}'")
		# si prefixe et suffixe valent True, la clef doit etre contenue
		else:
			self.cursor.execute(f"SELECT {self.primaryKey} FROM {self.tableName} WHERE {self.primaryKey} LIKE '%{key}%'")

		# resultat
		if self.cursor.fetchall() == []:
			return False
		else:
			return True

	# ajoute une nouvelle entrée dans la base de données
	def create(self, valeurs, lower):
		nomsColonnes = self.namesColonnes()
		# on vérifie que l'entrée n'existe pas déja
		if not self.verify(valeurs[self.primaryKeyIndex], False, False):
			text = f"INSERT INTO {self.tableName}("
			for k in nomsColonnes:
				text += f"{k},"
			text = f"{text[:-1]}) VALUES("
			for k in valeurs:
				if k == "NULL":
					text += "NULL,"
				elif lower:
					text += f"'{str(k).lower()}',"
				else:
					text += f"'{k}',"
			text = f"{text[:-1]})"
			try:
				self.cursor.execute(text)
			except sqlite3.OperationalError as e:
				raise Exit(f"[!] erreur dans l'opération : {e}")
		else:
			raise Exit(f"[!] {self.primaryKey} = {valeurs[self.primaryKeyIndex]}, cette entrée existe déjà")

	# supprime une entrée en la selectionnant avec la clef primaire
	def delete(self, key):
		# on vérifie que l'entrée existe
		if self.verify(key, False, False):
			self.cursor.execute(f"DELETE FROM {self.tableName} WHERE {self.primaryKey}='{key}'")
		else:
			raise Exit(f"[!] {self.primaryKey} = {key}, pas d'entrée corespondante")

	# modifie une entrée en la selectionnant avec la clef primaire (dans le champ valeurs)
	def modify(self, valeurs, lower):
		nomsColonnes = self.namesColonnes()
		# on vérifie que l'entrée existe
		if self.verify(valeurs[self.primaryKeyIndex], False, False):
			text = f"UPDATE {self.tableName} SET"
			for k in range(len(nomsColonnes)):
				if lower:
					text += f" {nomsColonnes[k]}='{str(valeurs[k]).lower()}',"
				else:
					text += f" {nomsColonnes[k]}='{valeurs[k]}',"
			text = f"{text[:-1]} WHERE {self.primaryKey} = '{valeurs[self.primaryKeyIndex]}'"
			try:
				self.cursor.execute(text)
			except sqlite3.OperationalError as e:
				raise Exit(f"[!] erreur dans l'opération : {e}")
		else:
			raise Exit(f"[!] {self.primaryKey} = {valeurs[self.primaryKeyIndex]}, pas d'entrée correspondante")

	# sauvegarde la base de donnée
	def save(self):
		self.connection.commit()

	# ferme la base de donnée
	def close(self):
		self.cursor.close()
		self.connection.close()

# utilise la classe obj_bdd pour lancer la methode : obj_bdd.commande(args)
def interact_bdd(path, tableName, commande, args=None, lower=True):
	s = None
	bdd = obj_bdd(path, tableName)

	if commande == "getDatas":
		if args is None or type(args) != str:
			raise Exit("[!] la commande 'getDatas' a besoin d'une clef primaire en argument")
		s = bdd.getDatas(args)

	elif commande == "verify":
		if args is None or not type(args) in [list,tuple] or len(args) != 3:
			raise Exit("[!] la commande 'verify' a besoin en argument de [clef primaire, prefixe(True/False), suffixe(True/False)]")
		s = bdd.verify(args[0], args[1], args[2])

	elif commande == "create":
		if args is None or not type(args) in [list,tuple] or len(args) != len(bdd.namesColonnes()):
			raise Exit(f"[!] la commande 'create' a besoin en arguments de {bdd.namesColonnes()}\nargument optionnel: lower [True/False]")
		bdd.create(args, lower)
		bdd.save()

	elif commande == "delete":
		if args is None or type(args) != str:
			raise Exit("[!] la commande 'delete' a besoin d'une clef primaire en argument")
		bdd.delete(args)
		bdd.save()

	elif commande == "modify":
		if args is None or not type(args) in [list,tuple] or len(args) != len(bdd.namesColonnes()):
			raise Exit(f"[!] la commande 'modify' a besoin en arguments de {bdd.namesColonnes()}argument optionnel: lower [True/False]")
		bdd.modify(args, lower)
		bdd.save()

	else:
		raise Exit(f"[!] la commande {commande} n'existe pas\nLes commandes possibles sont : \n verify\n getDatas\n getDatasSimilars\n getDatasInterval\n create\n delete\n modify\n passCommand\n")

	bdd.close()
	return s


## ~~~~~~~~~~~~~~~~~~~~~~~~~~	FILTRE MESSAGE PERSO   ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# classe de filtres personalisés
class filtres_perso:
	# detection des dates possibles
	class _date(MessageFilter):
		def filter(self, message):
			if message.text:
				try:
					temp = dt.strptime(message.text, "%d %m %Y")
					return True
				except ValueError:
					return False
	date = _date()

	# detection des heures possibles
	class _heure(MessageFilter):
		def filter(self, message):
			if message.text:
				try:
					temp = dt.strptime(message.text, "%H %M")
					return True
				except ValueError:
					return False
	heure = _heure()

	# detection des agences possibles
	class _agence(MessageFilter):
		def filter(self, message):
			if message.text:
				if message.text in ["appel medical", "adecco"]:
					return True
				return False
	agence = _agence()

# renvois sous forme lisible une ligne de la base de donnée
def bdd_to_string(extrait, mode="normal"):
	# si mode normal
	if mode == "normal":
		msg = " - ({}) {} à {}, de {} à {}".format(
			extrait[0],
			dt.strptime(extrait[1], "%Y/%m/%d").strftime("%a %-d %B"),
			extrait[2],
			extrait[3],
			extrait[4])
	# si mode récapitulatif
	elif mode == "recapitulatif":
		msg = "({}) {} à {}, de {} à {}".format(
			extrait[0],
			dt.strptime(extrait[1], "%Y/%m/%d").strftime("%a %-d %B"),
			extrait[2],
			extrait[3],
			extrait[4])
	# si mode mail
	elif mode == "mail":
		msg = "- {} à {}, de {} à {}".format(
			dt.strptime(extrait[1], "%Y/%m/%d").strftime("%a %-d %B"),
			extrait[2],
			extrait[3],
			extrait[4])
	# si mode raccourci
	elif mode == "court":
		msg = "{} à {}".format(
			dt.strptime(extrait[1], "%Y/%m/%d").strftime("%a %-d %B"),
			extrait[2])
	# si mode affichant seulement la clef primaire
	elif mode == "id":
		msg = extrait[1]

	else:
		print("[!] mode inconnu d'afficahge", mode)

	return msg

## ~~~~~~~~~~~~~~~~~~~~~~~~~~	   COMMANDES BOT	   ~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# fonction lancée par la commande '/start'
def start(update, context):
	update.message.reply_text("Coucou !\nAppuis sur '/' pour voir les commandes disponibles")

# les fonctions de la conversations de nouvelle missions
class conv_nouvelleMission():
	# conversation de nouvelle mission, commande n°1 de lancement et demande de la date
	def f_new_agence(update, context):
		# le clavier qu'on va renvoyer
		keyboard = [["adecco", "appel medical"]]
		# charge le clavier et l'envois
		update.message.reply_text(
			"Début de l'enregistrement d'une nouvelle mission\nentre '/stop' pour annuler à tout moment\n\n"
			"Avec quelle agence était la mission ?",
			reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
		)
		# renvoit l'étape suivante
		return AGENCE

	# conversation de nouvelle mission, commande n°1 de lancement et demande de la date
	def f_agence_date(update, context):
		global TO_SAVE
		# enregistrement de l'agence
		TO_SAVE.append(update.message.text)
		# le clavier qu'on va renvoyer
		##keyboard = [["aujourd'hui", "autre"]]
		# charge le clavier et l'envois
		update.message.reply_text(
			"ok, la date (en format 'JJ MM AAAA') de ta mission ?",
			reply_markup=ReplyKeyboardRemove()
			##reply_markup=ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
		)
		# renvoit l'étape suivante
		return DATE

	# conversation de nouvelle mission, commande n°2 pour enregistrer la date et demander le lieu
	def f_date_lieu(update, context):
		global TO_SAVE
		# enregistrement de la date
		TO_SAVE.append(dt.strptime(update.message.text, "%d %m %Y").strftime("%Y/%m/%d"))
		# la question suivante
		update.message.reply_text("ok, maintenant le lieu ?")
		# renvoit l'étape suivante
		return LIEU

	# conversation de nouvelle mission, commande n°3 pour enregistrer le lieu et demander l'heure de début
	def f_lieu_hDebut(update, context):
		global TO_SAVE
		# enregistrement du lieu
		TO_SAVE.append(update.message.text)
		# la question suivante
		update.message.reply_text("l'heure réelle (en format 'HH MM') de début de mission ?")
		# renvoit l'étape suivante
		return H_DEBUT

	# conversation de nouvelle mission, commande n°4 pour enregistrer l'heure de début et demander l'heure de fin
	def f_hDebut_hFin(update, context):
		global TO_SAVE
		# enregistrement de l'heure de début
		TO_SAVE.append(dt.strptime(update.message.text, "%H %M").strftime("%H:%M"))
		# la question suivante
		update.message.reply_text("l'heure réelle (en format 'HH MM') de fin de mission ?")
		# renvoit l'étape suivante
		return H_FIN

	# conversation de nouvelle mission, commande n°5 pour enregistrer l'heure de fin et clôturer
	def f_hFin_sauvegarde(update, context):
		global TO_SAVE
		# enregistrement de l'heure de fin
		TO_SAVE.append(dt.strptime(update.message.text, "%H %M").strftime("%H:%M"))
		# un petit récapitulatif
		update.message.reply_text(f"récapitulatif :\n{bdd_to_string(TO_SAVE, "recapitulatif")}")
		# sauvegarde de ces informations dans la base de donnée
		try:
			interact_bdd(BDD_PATH, BDD_TABLE, "create", TO_SAVE)
			# réponse pour dire que tout va bien
			update.message.reply_text("ok c'est bien enregistré")
		except Exit as e:
			# réponse pour dire qu'il y a eu une erreur
			print(e) #update.message.reply_text(f"code d'erreur : {e}")

		# remise a zéro des constantes
		TO_SAVE = []
		# fin de la conversation
		return ConversationHandler.END

	# conversation de nouvelle mission, commande d'annulation '/stop'
	def f_stop(update, context):
		global TO_SAVE
		# remise a zéro des constantes
		TO_SAVE = []
		# message d'annulation
		update.message.reply_text(
			"annulation de l'enregistrement",
			reply_markup=ReplyKeyboardRemove()
		)
		# fin de la conversation
		return ConversationHandler.END

# affiche les missions enregistrées dans la base de donnée
def affiche_missions(update, context):
	# toutes les données de la table et tri chronologique
	temp = interact_bdd(BDD_PATH, BDD_TABLE, "getDatas", "all")
	# si la base de donnée n'est pas vide
	if len(temp) > 0:
		msg = "toutes les missions enregistrées :\n"
		# pour chaque élément on ajoute une ligne
		for k in temp:
			msg += bdd_to_string(k) + "\n"
		update.message.reply_text(msg[:-1])
	# sinon la bdd est vide
	else:
		update.message.reply_text("pas de mission enregistrées :(\nutilises la commande '/nouvelle_mission'")

# supprime une mission de la base de données avec un clavier Inline
def supprime_mission(update, context):
	# le clavuer inline qu'on va remplir
	keyboard = []
	# toutes les données de la table et tri chronologique
	temp = interact_bdd(BDD_PATH, BDD_TABLE, "getDatas", "all")
	# on les mets dans le lavier inline en colonne
	for k in range(len(temp)):
		keyboard.append([InlineKeyboardButton(bdd_to_string(temp[k], "court"), callback_data="s_"+bdd_to_string(temp[k], "id"))])

	# la ligne pour annuler
	keyboard.append([InlineKeyboardButton("annuler", callback_data="s_annuler")])
	# charge le clavier et l'envois
	update.message.reply_text("sélectionnes pour supprimer :", reply_markup=InlineKeyboardMarkup(keyboard))

# envoit une capture d'écran des missions effectué pour donner les horaires exactes à l'agence
def horaires_mail(update, context):
	# récupère les constantes dans le .env
	with open(os.path.join(BASEPATH, ".env"), "r") as f:
		txt = f.read()
	server_name = REGEX_SVR_NAME.findall(txt)[0][12:]
	server_port = REGEX_SVR_PORT.findall(txt)[0][12:]
	mail_from = REGEX_MAIL_FROM.findall(txt)[0][10:]
	mail_mdp = REGEX_MAIL_MDP.findall(txt)[0][9:]
	mail_to = REGEX_MAIL_TO.findall(txt)[0][8:]
	del txt
	# toutes les données de la table et tri chronologique
	temp = interact_bdd(BDD_PATH, BDD_TABLE, "getDatas", "all")
	# on crée les deux textes pour les deux agences
	text_adecco = ""
	text_appelMedical = ""
	# répartition des mission de la base de donnée en fonction des agences
	for k in temp:
		if k[0] == "appel medical":
			text_appelMedical += bdd_to_string(k, "mail") + "\n"
		elif k[0] == "adecco":
			text_adecco += bdd_to_string(k, "mail") + "\n"
		else:
			print("[!] agence inconnue :", k)

	# rédaction du message
	message = MIMEMultipart("alternative")
	message["Subject"] = "[INTERIM] horaires réelles de missions"
	message["From"] = mail_from
	message["To"] = mail_to

	texte = f"""
Missions avec Appel Medical :\n
Bonjour,\n
veuillez trouver ci-joint les horaires réelles des missions que j'ai effectuée :\n
{text_appelMedical}
Cordialement,

-------------------------

Missions avec adecco :\n
Bonjour,\n
veuillez trouver ci-joint les horaires réelles des missions que j'ai effectuée :\n
{text_adecco}
Cordialement,

-------------------------
bot TELEGRAM @missions_interim_bot
by mgl corp."""
	texte_mime = MIMEText(texte, "plain")
	message.attach(texte_mime)

	# connection au serveur
	try:
		with smtplib.SMTP(server_name, server_port) as server:
			server.starttls()
			server.login(mail_from, mail_mdp)
			server.sendmail(mail_from, mail_to, message.as_string())

		# réponse
		update.message.reply_text("mail envoyé")
	# si un problème
	except Exception as e:
		print(e)
		update.message.reply_text(f"erreur dans l'envoi du mail :\n{e}")
	finally:
		del server_name
		del server_port
		del mail_from
		del mail_mdp
		del mail_to

# exporte toutes les missions enregistrées dans un fichier excel
def exporte_excel(update, context):
	# nombre de lignes (toutes les données de la table)
	nb_lignes = len(interact_bdd(BDD_PATH, BDD_TABLE, "getDatas", "all"))
	# si la base de donnée est vide
	if nb_lignes == 0:
		update.message.reply_text("pas de mission enregistrées")
	# sinon on envoit tout
	else:
		# le clavier Inline de confirmation de l'action
		keyboard = [
			[InlineKeyboardButton("continuer", callback_data="e_continuer")],
			[InlineKeyboardButton("annuler", callback_data="e_annuler")]
		]
		# charge le clavier et l'envois
		update.message.reply_text("veux-tu créer le fichier Excel ?\nAttention, il sera impossible ensuite d'envoyer le mail", reply_markup=InlineKeyboardMarkup(keyboard))

# fonction lancée par un appuis le clavier inline
def button(update, context):
	query = update.callback_query
	# si la query commence par 's', on supprime l'entrée
	if query.data[:2] == "s_":
		query.answer()
		if query.data[2:] == "annuler":
			# réponse au client( change le message précédemment envoyé)
			query.edit_message_text(text="annulé")
		else:
			try:
				interact_bdd(BDD_PATH, BDD_TABLE, "delete", query.data[2:])
				# réponse au client (obligatoire sinon bug sur certains clients)
				query.edit_message_text(text="mission supprimée")
			except Exit as e:
				# réponse pour dire qu'il y a eu une erreur
				print(e) #query.edit_message_text(text=f"code d'erreur : {e}")

	# si la query commence par 'e', on exporte les missions
	elif query.data[:2] == "e_":
		query.answer()
		# si c'est le code d'annulation
		if query.data[2:] == "annuler":
			# réponse au client( change le message précédemment envoyé)
			query.edit_message_text(text="annulé")

		# si c'est le code de continuation
		if query.data[2:] == "continuer":
			# toutes les données de la table et tri chronologique
			temp = interact_bdd(BDD_PATH, BDD_TABLE, "getDatas", "all")
			# on crée les deux listes pour les deux agences
			list_adecco = []
			list_appelMedical = []
			# répartition des mission de la base de donnée en fonction des agences
			for k in temp:
				if k[0] == "appel medical":
					list_appelMedical.append(k)
				elif k[0] == "adecco":
					list_adecco.append(k)
				else:
					print("[!] agence inconnue :", k)
			# nombre de lignes qu'on va écrire
			nb_lignes_appelMedical = len(list_appelMedical)
			nb_lignes_adecco = len(list_adecco)
			# création du fichier excel temporaire
			wb = Workbook()
			ws = wb.active
			# missions de appel medical
			i = -1
			for row in ws.iter_rows(max_col=MAX_COL, min_row=0, max_row=nb_lignes_appelMedical):
				if i == -1:
					row[0].value = "appel medical"
				else:
					# mise en forme compréhensible par excel des données
					row[0].value = dt.strptime(list_appelMedical[i][1], "%Y/%m/%d").strftime("%d/%m/%Y")
					row[1].value = list_appelMedical[i][2]
					row[4].value = list_appelMedical[i][3]
					row[5].value = list_appelMedical[i][4]
				i += 1
			# mission de adecco
			i = -1
			for row in ws.iter_rows(max_col=MAX_COL, min_row=nb_lignes_appelMedical+3, max_row=nb_lignes_appelMedical+3+nb_lignes_adecco):
				if i == -1:
					row[0].value = "adecco"
				else:
					# mise en forme compréhensible par excel des données
					row[0].value = dt.strptime(list_adecco[i][1], "%Y/%m/%d").strftime("%d/%m/%Y")
					row[1].value = list_adecco[i][2]
					row[4].value = list_adecco[i][3]
					row[5].value = list_adecco[i][4]
				i += 1
			# le chemin temporaire du excel
			tempPathExcel = os.path.join(BASEPATH, "extrait.xlsx")
			# sauvegarde du excel
			wb.save(filename=tempPathExcel)
			# envoit du fichier
			query.edit_message_text("excel envoyé")
			context.bot.send_document(chat_id=query.message.chat_id, document=open(tempPathExcel, "rb"))
			# suppression du excel
			os.remove(tempPathExcel)
			# nettoyage de la base de données
			try:
				for k in temp:
					interact_bdd(BDD_PATH, BDD_TABLE, "delete", k[1]) # la clef primaire est en position 1
					print("deleted :", k[0])
				# envoi un nouveau message
				context.bot.send_message(chat_id=query.message.chat_id, text="base de donnée nettoyée")
			except Exit as e:
				# réponse pour dire qu'il y a eu une erreur
				print(e) #context.bot.send_message(chat_id=query.message.chat_id, text=f"code d'erreur : {e}")

# affiche l'aide
def help(update, context):
	update.message.reply_text("""\
Commandes disponibles:
/nouvelle_mission : enregistre une nouvelle mission
/affiche_missions : affiche toutes les missions
/supprime_mission : supprime une mission
/horaires_mail : envoie par mail les horaires réels pour l'agence
/exporte_excel : renvoit le fichier excel rempli
/help : affiche l'aide""")

# affiche les erreurs rencontrés par le programme
def error(update, context):
	print(f"Update '{update}' \ncaused error '{context.error}'")


## ~~~~~~~~~~~~~~~~~~~~~~~~~~	FONCTION PRINCIPALE	~~~~~~~~~~~~~~~~~~~~~~~~~~ ##

# la fonction principale du bot
def main():
	# récupere le token d'identitification dans le .env
	if os.path.isfile(os.path.join(BASEPATH, ".env")):
		with open(os.path.join(BASEPATH, ".env"), "r") as f:
			try:
				# création du bot avec son token d'authentification (retire le 'token=' du début)
				bot = Updater(REGEX_TOKEN.findall(f.read())[0][6:], use_context=True)
			except Exception as e:
				raise e
	else:
		raise Exit("[!] le fichier .env contenant le token d'identitification n'existe pas")
	# initialisation de la base de donnée (crée la base de donnée et la table si elle n'existe pas)
	DATABASE = obj_bdd(BDD_PATH, BDD_TABLE)
	# création du conversation handler pour créer un nouvel enregistrement
	conversation_nouvelleMission = ConversationHandler(
		entry_points=[CommandHandler("nouvelle_mission", conv_nouvelleMission.f_new_agence)],
		states={
			AGENCE: [MessageHandler(filtres_perso.agence, conv_nouvelleMission.f_agence_date)],
			DATE: [MessageHandler(filtres_perso.date, conv_nouvelleMission.f_date_lieu)],
			LIEU: [MessageHandler(Filters.text & ~Filters.command, conv_nouvelleMission.f_lieu_hDebut)],
			H_DEBUT: [MessageHandler(filtres_perso.heure, conv_nouvelleMission.f_hDebut_hFin)],
			H_FIN: [MessageHandler(filtres_perso.heure, conv_nouvelleMission.f_hFin_sauvegarde)],
		},
		fallbacks=[CommandHandler("stop", conv_nouvelleMission.f_stop)],
	)
	# ajout des gestionnaires de commande par ordre d'importance
	# la commande /start
	bot.dispatcher.add_handler(CommandHandler("start", start))
	# la commande de conversation /nouvelle_mission
	bot.dispatcher.add_handler(conversation_nouvelleMission)
	# la commande /affiche_missions
	bot.dispatcher.add_handler(CommandHandler("affiche_missions", affiche_missions))
	# la commande /supprime_mission
	bot.dispatcher.add_handler(CommandHandler("supprime_mission", supprime_mission))
	# la commande /horaires_mail
	bot.dispatcher.add_handler(CommandHandler("horaires_mail", horaires_mail))
	# la commande /exporte_excel
	bot.dispatcher.add_handler(CommandHandler("exporte_excel", exporte_excel))
	# le clavier inline
	bot.dispatcher.add_handler(CallbackQueryHandler(button))
	# la commande /help
	bot.dispatcher.add_handler(CommandHandler("help", help))
	# gestion des erreurs
	bot.dispatcher.add_error_handler(error)

	# lance le bot
	bot.start_polling()
	# continue le programme jusqu'à la reception d'un signal de fin (par ex: CTRL-C)
	bot.idle()

# lance la fonction principale
if __name__ == "__main__":
	main()


# fin
